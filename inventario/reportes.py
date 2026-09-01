# inventario/reportes.py
"""
Generación de reportes Excel (openpyxl) para la auditoría. Usa
`inventario.services` como única fuente de verdad de los cálculos: este
módulo solo formatea y arma el libro, nunca recalcula cifras por su cuenta.

Nota (prompt de rediseño): el Excel se simplificó a solo 2 hojas
(Resumen, Movimientos). Antes incluía hojas de gráficos, pero esas
gráficas ahora se ven directamente en la app (pantalla de Reportes), así
que duplicarlas en el Excel ya no aportaba y solo lo hacía más pesado de
revisar para un usuario no técnico.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .services import movimientos_periodo, resumen_general

FORMATO_MONEDA = '"Q"#,##0.00'
FORMATO_FECHA = "DD/MM/YYYY"

# (encabezado de columna, clave en el dict de resumen_producto/resumen_general)
COLUMNAS_RESUMEN = [
    ("Producto", "producto_nombre"),
    ("Categoría", "categoria"),
    ("Unidades compradas", "unidades_compradas"),
    ("Invertido", "invertido"),
    ("Unidades vendidas", "unidades_vendidas"),
    ("Ingreso", "ingreso"),
    ("Costo de lo vendido", "costo_de_lo_vendido"),
    ("Ganancia bruta", "ganancia_bruta"),
    ("Unidades merma", "unidades_merma"),
    ("Pérdida por merma", "perdida_por_merma"),
    # Ajustes de inventario confirmados en el período (prompt 34). Con
    # signo: negativo = faltante encontrado al contar, positivo =
    # sobrante. El resumen ya calculaba este número pero no lo mostraba
    # en ningún lado, así que un faltante confirmado —una pérdida real de
    # producto— no aparecía en el reporte por ninguna parte. No entra en
    # la ganancia neta a propósito: un ajuste corrige un error de
    # registro, y darlo por pérdida sin más mezclaría dos cosas
    # distintas. Queda como columna propia para que se vea y se decida.
    ("Unidades ajuste", "unidades_ajuste"),
    ("Ganancia neta", "ganancia_neta"),
    ("Stock teórico al cierre", "stock_teorico_al_cierre"),
]
_COLUMNAS_MONETARIAS = {
    "Invertido",
    "Ingreso",
    "Costo de lo vendido",
    "Ganancia bruta",
    "Pérdida por merma",
    "Ganancia neta",
}


def _hoja_resumen(wb, fecha_inicio, fecha_fin, productos):
    resumen = resumen_general(fecha_inicio, fecha_fin, productos=productos)
    ws = wb.active
    ws.title = "Resumen"

    ws.append([nombre for nombre, _ in COLUMNAS_RESUMEN])
    for celda in ws[1]:
        celda.font = Font(bold=True)

    for fila in resumen["productos"]:
        ws.append([fila[clave] for _, clave in COLUMNAS_RESUMEN])

    fila_totales_idx = ws.max_row + 1
    totales = resumen["totales"]
    # "Stock teórico al cierre" no tiene un total con sentido (sumar stock
    # de productos distintos no es una cifra útil), queda en blanco.
    valores_totales = ["TOTAL", ""] + [totales.get(clave, "") for _, clave in COLUMNAS_RESUMEN[2:]]
    ws.append(valores_totales)
    for celda in ws[fila_totales_idx]:
        celda.font = Font(bold=True)

    for col_idx, (nombre, _) in enumerate(COLUMNAS_RESUMEN, start=1):
        letra = get_column_letter(col_idx)
        ws.column_dimensions[letra].width = max(len(nombre) + 2, 14)
        if nombre in _COLUMNAS_MONETARIAS:
            for fila_idx in range(2, fila_totales_idx + 1):
                ws[f"{letra}{fila_idx}"].number_format = FORMATO_MONEDA

    return ws, fila_totales_idx


def _hoja_movimientos(wb, fecha_inicio, fecha_fin, productos):
    ws = wb.create_sheet("Movimientos")
    encabezados = ["Fecha", "Producto", "Tipo", "Cantidad", "Precio o costo unitario", "Usuario", "Motivo/Notas"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    filas = movimientos_periodo(fecha_inicio, fecha_fin, productos=productos, descendente=False)
    for fila in filas:
        ws.append(
            [
                fila["fecha"],
                str(fila["producto"]),
                fila["tipo"],
                fila["cantidad"],
                fila["valor_unitario"],
                fila["usuario"],
                fila["detalle"],
            ]
        )

    for fila_idx in range(2, ws.max_row + 1):
        ws[f"A{fila_idx}"].number_format = FORMATO_FECHA
        ws[f"E{fila_idx}"].number_format = FORMATO_MONEDA

    for col_idx, ancho in enumerate([14, 22, 20, 10, 22, 14, 40], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    return ws


def generar_reporte_excel(fecha_inicio, fecha_fin, productos=None) -> BytesIO:
    """
    Construye el libro de auditoría para [fecha_inicio, fecha_fin] (ambos
    extremos inclusive) y lo devuelve como BytesIO listo para servir o
    guardar. `productos`: iterable de Producto, o None para todos los
    productos activos (ver resumen_general). Dos hojas: Resumen y
    Movimientos — las gráficas se consultan en pantalla, en la app.
    """
    wb = Workbook()
    _hoja_resumen(wb, fecha_inicio, fecha_fin, productos)
    _hoja_movimientos(wb, fecha_inicio, fecha_fin, productos)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
